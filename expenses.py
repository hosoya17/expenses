import tkinter as tk
import tkinter.ttk as ttk
import openpyxl
from tkcalendar import Calendar
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

class Application(tk.Frame):
    def __init__(self, root):
        super().__init__(root, width=562, height=500, borderwidth=1, relief='groove')
        self.root = root
        self.pack()
        self.pack_propagate(False)
        self.create_widgets()
        self.error_message = None
        self.last_error_message = None 

    def create_widgets(self):
        title_label = tk.Label(self, text='家計簿アプリ', font=('Yu Gothic UI', '24', 'normal'))
        title_label.pack(pady=10, anchor='w')

        self.calender = Calendar(self, locale='ja_JP', showweeknumbers=False)
        self.calender.pack(expand=True, fill=tk.BOTH)

        self.add_btn = tk.Button(self, text='追加', width=10, height=2)
        self.add_btn['command'] = self.add
        self.add_btn.pack(pady=10, anchor='s')

    def add(self):
        self.selected_date = self.calender.selection_get()
        self.selected_date = self.selected_date.strftime('%Y年%m月%d日')

        self.add_window = tk.Toplevel(self.root)
        self.add_window.title('データ追加')
        self.add_window.geometry('300x350')

        title_label = tk.Label(self.add_window, text=self.selected_date, font=('Yu Gothic UI', '20', 'normal'))
        title_label.pack()

        money_label = tk.Label(self.add_window, text='金額', font=('Yu Gothic UI', '15', 'normal'))
        money_label.place(x=0, y=70)

        unit_label = tk.Label(self.add_window, text='円', font=('Yu Gothic UI', '10', 'normal'))
        unit_label.place(x=213, y=80)

        category_label = tk.Label(self.add_window, text='カテゴリ', font=('Yu Gothic UI', '15', 'normal'))
        category_label.place(x=0, y=110)

        payment_label = tk.Label(self.add_window, text='支払方法', font=('Yu Gothic UI', '15', 'normal'))
        payment_label.place(x=0, y=148)

        self.money_entry = tk.Entry(self.add_window)
        self.money_entry.place(x=90, y=80)

        self.category_list = ttk.Combobox(self.add_window, values=['食費', '日用品', '交通費', '水道・光熱', '通信費', 'お住まい', '衣類', '自動車', 'レジャー', '美容', '交際費', '書籍', '保険', '税金', '医療費', '振替手数料'], state='readonly', width=15)
        self.category_list.place(x=90, y=118)

        self.payment_list = ttk.Combobox(self.add_window, values=['現金', 'クレジットカード', 'QRコード決済', '交通系IC', '銀行'], state='readonly', width=15)
        self.payment_list.place(x=90, y=157)

        data_add_btn = tk.Button(self.add_window, text='追加', width=10)
        data_add_btn['command'] = self.data_add
        data_add_btn.place(x=110, y=310)

        self.add_window.selected_radio = tk.StringVar(value='支出')
        self.radio_1 = tk.Radiobutton(self.add_window, text='支出', value='支出', variable=self.add_window.selected_radio, command=self.on_radio_select)
        self.radio_2 = tk.Radiobutton(self.add_window, text='収入', value='収入', variable=self.add_window.selected_radio, command=self.on_radio_select)
        self.radio_1.place(x=0, y=45)
        self.radio_2.place(x=50, y=45)

        self.payment_label = payment_label
        self.payment_list = self.payment_list

    def data_add(self):
        selected_radio = self.add_window.selected_radio.get()
        input_money = self.money_entry.get()
        selected_category_list = self.category_list.get()
        selected_payment_list = self.payment_list.get()
        self.remove_error_message()

        if selected_radio == '支出':
            if(len(str(input_money)) == 0 or len(selected_category_list) == 0 or len(selected_payment_list) == 0):
                error_message = tk.Label(self.add_window, text='全ての項目を入力してください。', font=('Yu Gothic UI', '8', 'normal'), foreground='#ff0000')
                error_message.place(x=0, y=180)
                self.last_error_message = error_message
            else:
                self.success_window = tk.Toplevel(self.root)
                self.success_window.title('成功画面')
                self.success_window.geometry('250x150')

                success_label = tk.Label(self.success_window, text='データを追加しました。')
                success_label.place(x=70, y=50)

                back_btn = tk.Button(self.success_window, text='閉じる', width=10)
                back_btn['command'] = self.success_window_close
                back_btn.place(x=70, y=110)

                self.excel()

                self.success_window.protocol("WM_DELETE_WINDOW", self.success_window_close)
        else:
            if(len(str(input_money)) == 0 or len(selected_category_list) == 0):
                error_message = tk.Label(self.add_window, text='全ての項目を入力してください。', font=('Yu Gothic UI', '8', 'normal'), foreground='#ff0000')
                error_message.place(x=0, y=140)
                self.last_error_message = error_message
            else:
                self.success_window = tk.Toplevel(self.root)
                self.success_window.title('成功画面')
                self.success_window.geometry('250x150')

                success_label = tk.Label(self.success_window, text='データを追加しました。')
                success_label.place(x=70, y=50)

                back_btn = tk.Button(self.success_window, text='閉じる', width=10)
                back_btn['command'] = self.success_window_close
                back_btn.place(x=70, y=110)

                self.excel()

                self.success_window.protocol("WM_DELETE_WINDOW", self.success_window_close)

    def on_radio_select(self):
        selected_radio = self.add_window.selected_radio.get()
        if(selected_radio == '支出'):
            self.payment_label.place(x=0, y=148)
            self.payment_list.place(x=90, y=157)
        else:
            self.category_list = ttk.Combobox(self.add_window, values=['給料', '臨時収入'], state='readonly', width=15)
            self.category_list.place(x=90, y=118)
            self.payment_label.place_forget()
            self.payment_list.place_forget()

    def success_window_close(self):
        if(self.last_error_message):
            self.last_error_message.place(x=0, y=180)
        self.success_window.destroy()

    def remove_error_message(self):
        if(self.last_error_message):
            self.last_error_message.destroy()
            self.last_error_message = None

    def create_excel_file(self, file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    def excel(self):
        wb = load_workbook('C:\\Python\\expenses\\Excel\\家計簿.xlsx')
        sheet = wb.active
        row = sheet.max_row + 1

        selected_date = self.selected_date
        money_value = self.money_entry.get()
        category_value = self.category_list.get()
        payment_value = self.payment_list.get()
        selected_radio = self.add_window.selected_radio.get()
        if(selected_radio == '支出'):
            money_value = int(money_value) * -1

        sheet.cell(row=row, column=1, value=selected_date).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet.cell(row=row, column=2, value=int(money_value)).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet.cell(row=row, column=3, value=payment_value).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet.cell(row=row, column=4, value=category_value).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet.cell(row=row, column=5, value=selected_radio).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        wb.save('C:\\Python\\expenses\\Excel\\家計簿.xlsx')
        wb.close()


root = tk.Tk()
root.title('家計簿アプリ')
root.geometry('562x500')
app = Application(root=root)
root.mainloop()
